from __future__ import annotations

from openpyxl import Workbook
import pytest

from app.agent.builtin_plugins.documents.rendering.xlsx_formula import (
    evaluate_workbook_formulas,
    format_computed_value,
)


def test_evaluates_cross_sheet_dependencies_ranges_and_core_functions() -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Assumptions"
    inputs["B1"] = 100
    inputs["B2"] = 0.1
    forecast = workbook.create_sheet("Forecast")
    forecast["B1"] = "='Assumptions'!B1"
    forecast["B2"] = "=B1*(1+'Assumptions'!B2)"
    forecast["B3"] = "=B2*(1+'Assumptions'!B2)"
    dashboard = workbook.create_sheet("Dashboard")
    dashboard["B1"] = "=SUM('Forecast'!B1:B3)"
    dashboard["B2"] = "=AVERAGE('Forecast'!B1:B3)"
    dashboard["B3"] = "=MAX('Forecast'!B1:B3)"
    dashboard["B4"] = '=IF(B3>120,"ready","wait")'

    evaluation = evaluate_workbook_formulas(workbook)

    assert evaluation.formula_count == 7
    assert evaluation.evaluated_count == 7
    assert evaluation.issues == []
    assert evaluation.values[("Forecast", "B3")] == pytest.approx(121)
    assert evaluation.values[("Dashboard", "B1")] == pytest.approx(331)
    assert round(evaluation.values[("Dashboard", "B2")], 4) == 110.3333
    assert evaluation.values[("Dashboard", "B4")] == "ready"
    assert format_computed_value(0.125, "0.0%") == "12.5%"


def test_reports_unsupported_functions_and_cycles_without_aborting_scan() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "=A2"
    sheet["A2"] = "=A1"
    sheet["B1"] = "=XLOOKUP(1,C1:C2,D1:D2)"
    sheet["C1"] = 1
    sheet["D1"] = 2

    evaluation = evaluate_workbook_formulas(workbook)

    assert evaluation.formula_count == 3
    assert evaluation.evaluated_count == 0
    assert {issue["code"] for issue in evaluation.issues} == {
        "formula-cycle",
        "formula-evaluation-failed",
    }

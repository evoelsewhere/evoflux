"""Editable-first XLSX authoring through OpenXML and ``openpyxl``.

The pipeline supports both net-new workbooks and targeted edits to uploaded
XLSX templates.  Existing workbooks are imported before modification, values
and formulas are edited without touching formatting unless explicitly asked,
and every worksheet is rendered and formula-scanned before publication.
"""

from __future__ import annotations

import asyncio
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
import json
import math
from pathlib import Path
import re
from typing import Any, Literal

from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator

from app.services.office.internal_rendering import render_xlsx_workbook
from app.services.office.runtime import file_sha256

# Backward-compatible public name retained for callers outside this module.
xlsx_sha256 = file_sha256


_A1_RANGE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*(?::[A-Z]{1,3}[1-9][0-9]*)?$")
_FORMULA_ERRORS = r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A|#NUM!|#NULL!"
MAX_SHEETS = 50
MAX_OPERATIONS = 1000


JsonScalar = str | int | float | bool | None


class RangeFormat(BaseModel):
    model_config = ConfigDict(extra="forbid")

    fill: str | None = Field(default=None, pattern=r"^#[0-9A-Fa-f]{6}$")
    font: dict[str, Any] | None = None
    borders: dict[str, Any] | None = None
    number_format: str | None = Field(default=None, max_length=120)
    horizontal_alignment: Literal["left", "center", "right", "general"] | None = None
    vertical_alignment: Literal["top", "center", "bottom"] | None = None
    wrap_text: bool | None = None
    column_width: float | None = Field(default=None, ge=2, le=120)
    row_height: float | None = Field(default=None, ge=8, le=300)


class WriteRange(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["write_range"] = "write_range"
    range: str
    values: list[list[JsonScalar]] | None = None
    formulas: list[list[str | None]] | None = None
    dates: list[list[str | None]] | None = None
    format: RangeFormat | None = None

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("range must use bounded A1 notation")
        return value.upper()

    @model_validator(mode="after")
    def validate_payload(self) -> WriteRange:
        supplied = sum(
            item is not None for item in (self.values, self.formulas, self.dates)
        )
        if supplied != 1:
            raise ValueError(
                "write_range requires exactly one of values, formulas, or dates"
            )
        if self.formulas is not None:
            for row in self.formulas:
                for formula in row:
                    if formula is not None and not formula.startswith("="):
                        raise ValueError("every formula must start with '='")
        return self


class StyleRange(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["style_range"] = "style_range"
    range: str
    format: RangeFormat

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("range must use bounded A1 notation")
        return value.upper()


class MergeRange(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["merge", "unmerge"]
    range: str

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if ":" not in value or not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("merge range must be a multi-cell A1 range")
        return value.upper()


class ClearRange(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["clear"] = "clear"
    range: str
    apply_to: Literal["contents", "formats", "all"] = "contents"

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("range must use bounded A1 notation")
        return value.upper()


class FreezePanes(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["freeze_panes"] = "freeze_panes"
    rows: int = Field(default=0, ge=0, le=100)
    columns: int = Field(default=0, ge=0, le=50)


class DataValidation(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["data_validation"] = "data_validation"
    range: str
    values: list[str] = Field(min_length=1, max_length=200)

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("range must use bounded A1 notation")
        return value.upper()


class ConditionalFormat(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["conditional_format"] = "conditional_format"
    range: str
    rule_type: Literal["cellIs", "colorScale", "dataBar", "containsText", "expression"]
    config: dict[str, Any]

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("range must use bounded A1 notation")
        return value.upper()


class AddTable(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["add_table"] = "add_table"
    range: str
    name: str = Field(pattern=r"^[A-Za-z_][A-Za-z0-9_]{0,254}$")
    has_headers: bool = True
    style: str | None = Field(default=None, max_length=100)

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if ":" not in value or not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("table range must be a multi-cell A1 range")
        return value.upper()


class AddChart(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["add_chart"] = "add_chart"
    chart_type: Literal["bar", "line", "area", "pie", "doughnut", "scatter"]
    source_range: str
    start_cell: str
    end_cell: str
    title: str = Field(min_length=1, max_length=200)
    has_legend: bool = True
    y_number_format: str | None = Field(default=None, max_length=120)

    @field_validator("source_range", "start_cell", "end_cell")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("chart references must use bounded A1 notation")
        return value.upper()


class DeleteDrawings(BaseModel):
    model_config = ConfigDict(extra="forbid")

    operation: Literal["delete_drawings"] = "delete_drawings"


class AutofitRange(BaseModel):
    """Size columns or rows from their rendered content.

    Preferred over guessing ``column_width``: a numeric cell narrower than its
    formatted text renders as ``#####`` in Excel, which the fit check reports.
    """

    model_config = ConfigDict(extra="forbid")

    operation: Literal["autofit_columns", "autofit_rows"]
    range: str

    @field_validator("range")
    @classmethod
    def validate_range(cls, value: str) -> str:
        if not _A1_RANGE.fullmatch(value.upper()):
            raise ValueError("range must use bounded A1 notation")
        return value.upper()


SheetOperation = (
    WriteRange
    | StyleRange
    | MergeRange
    | ClearRange
    | FreezePanes
    | DataValidation
    | ConditionalFormat
    | AddTable
    | AddChart
    | DeleteDrawings
    | AutofitRange
)


class WorksheetPlan(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str = Field(min_length=1, max_length=31)
    create_if_missing: bool = False
    show_grid_lines: bool | None = None
    operations: list[SheetOperation] = Field(
        default_factory=list, max_length=MAX_OPERATIONS
    )


class WorkbookProject(BaseModel):
    model_config = ConfigDict(extra="forbid")

    schema_version: Literal[1] = 1
    title: str = Field(min_length=1, max_length=240)
    mode: Literal["new", "template"]
    source_sha256: str | None = Field(default=None, pattern=r"^[0-9a-f]{64}$")
    template_confirmed: bool = False
    sheets: list[WorksheetPlan] = Field(min_length=1, max_length=MAX_SHEETS)

    @model_validator(mode="after")
    def validate_mode(self) -> WorkbookProject:
        names = [sheet.name.casefold() for sheet in self.sheets]
        if len(names) != len(set(names)):
            raise ValueError("worksheet plan names must be unique")
        if self.mode == "template":
            if not self.template_confirmed or self.source_sha256 is None:
                raise ValueError(
                    "template mode requires template_confirmed=true and source_sha256"
                )
        elif self.source_sha256 is not None or self.template_confirmed:
            raise ValueError("new mode must not declare template lineage")
        return self


@dataclass
class XlsxPipelineResult:
    action: str
    work_dir: Path
    source_xlsx: Path | None = None
    output: Path | None = None
    manifest_path: Path | None = None
    previews: list[Path] = field(default_factory=list)
    issues: list[dict[str, Any]] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return not any(issue.get("severity") == "error" for issue in self.issues)

    def to_dict(self) -> dict[str, Any]:
        return {
            "action": self.action,
            "work_dir": str(self.work_dir),
            "source_xlsx": str(self.source_xlsx) if self.source_xlsx else None,
            "output": str(self.output) if self.output else None,
            "manifest_path": str(self.manifest_path) if self.manifest_path else None,
            "previews": [str(path) for path in self.previews],
            "issues": self.issues,
            "passed": self.passed,
            **self.metadata,
        }


def load_workbook_project(path: Path) -> WorkbookProject:
    return WorkbookProject.model_validate_json(path.read_text(encoding="utf-8"))


def workbook_catalog() -> dict[str, Any]:
    return {
        "workflow": "editable-openxml-xlsx",
        "invariants": [
            "Use typed OpenXML operations for every workbook write and export.",
            "Render and inspect an uploaded workbook before template edits.",
            "Do not overwrite cell formatting when only values or formulas change.",
            "Keep inputs typed and derived values formula-driven.",
            "Scan formula errors and render every worksheet before publishing.",
            "Size columns with autofit_columns instead of guessing column_width.",
            "Never overwrite the uploaded source workbook.",
        ],
        "operations": [
            "write_range",
            "style_range",
            "merge",
            "unmerge",
            "clear",
            "freeze_panes",
            "data_validation",
            "conditional_format",
            "add_table",
            "add_chart",
            "delete_drawings",
            "autofit_columns",
            "autofit_rows",
        ],
        "project_json_schema": WorkbookProject.model_json_schema(),
    }


def _hex(value: str | None) -> str | None:
    return value.lstrip("#") if value else None


def _iter_cells(sheet: Any, range_address: str) -> list[Any]:
    value = sheet[range_address]
    if not isinstance(value, tuple):
        return [value]
    cells: list[Any] = []
    for row in value:
        cells.extend(row if isinstance(row, tuple) else (row,))
    return cells


def _side(value: Any) -> Any:
    from openpyxl.styles import Side

    if isinstance(value, str):
        return Side(style=value)
    if isinstance(value, dict):
        return Side(
            style=value.get("style") or value.get("line_style") or "thin",
            color=_hex(value.get("color")),
        )
    return Side()


def _apply_format(sheet: Any, range_address: str, value: RangeFormat) -> None:
    from openpyxl.styles import Border, Font, PatternFill

    for cell in _iter_cells(sheet, range_address):
        if value.fill is not None:
            cell.fill = PatternFill("solid", fgColor=_hex(value.fill))
        if value.font is not None:
            fields = {
                key: item
                for key, item in value.font.items()
                if key
                in {
                    "name",
                    "size",
                    "bold",
                    "italic",
                    "underline",
                    "strike",
                    "vertAlign",
                }
            }
            if "size" in fields:
                fields["sz"] = fields.pop("size")
            if color := value.font.get("color"):
                fields["color"] = _hex(str(color))
            cell.font = Font(**fields)
        if value.borders is not None:
            cell.border = Border(
                left=_side(value.borders.get("left")),
                right=_side(value.borders.get("right")),
                top=_side(value.borders.get("top")),
                bottom=_side(value.borders.get("bottom")),
            )
        if value.number_format is not None:
            cell.number_format = value.number_format
        alignment = copy(cell.alignment)
        if value.horizontal_alignment is not None:
            alignment.horizontal = value.horizontal_alignment
        if value.vertical_alignment is not None:
            alignment.vertical = value.vertical_alignment
        if value.wrap_text is not None:
            alignment.wrap_text = value.wrap_text
        cell.alignment = alignment
    from openpyxl.utils.cell import range_boundaries

    min_column, min_row, max_column, max_row = range_boundaries(range_address)
    if value.column_width is not None:
        from openpyxl.utils import get_column_letter

        for column in range(min_column, max_column + 1):
            sheet.column_dimensions[
                get_column_letter(column)
            ].width = value.column_width
    if value.row_height is not None:
        for row in range(min_row, max_row + 1):
            sheet.row_dimensions[row].height = value.row_height


def _write_matrix(sheet: Any, range_address: str, matrix: list[list[Any]]) -> None:
    from openpyxl.utils.cell import range_boundaries

    min_column, min_row, max_column, max_row = range_boundaries(range_address)
    expected_rows = max_row - min_row + 1
    expected_columns = max_column - min_column + 1
    if len(matrix) != expected_rows or any(
        len(row) != expected_columns for row in matrix
    ):
        raise ValueError(
            f"matrix dimensions do not match {range_address}: "
            f"expected {expected_rows}x{expected_columns}"
        )
    for row_offset, row in enumerate(matrix):
        for column_offset, item in enumerate(row):
            sheet.cell(min_row + row_offset, min_column + column_offset).value = item


def _add_chart(sheet: Any, operation: AddChart) -> None:
    from openpyxl.chart import (
        AreaChart,
        BarChart,
        DoughnutChart,
        LineChart,
        PieChart,
        Reference,
        ScatterChart,
    )
    from openpyxl.utils.cell import range_boundaries

    chart_types = {
        "bar": BarChart,
        "line": LineChart,
        "area": AreaChart,
        "pie": PieChart,
        "doughnut": DoughnutChart,
        "scatter": ScatterChart,
    }
    chart: Any = chart_types[operation.chart_type]()
    min_column, min_row, max_column, max_row = range_boundaries(operation.source_range)
    if max_column > min_column:
        data = Reference(
            sheet,
            min_col=min_column + 1,
            max_col=max_column,
            min_row=min_row,
            max_row=max_row,
        )
        categories = Reference(
            sheet,
            min_col=min_column,
            min_row=min_row + 1,
            max_row=max_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
    else:
        data = Reference(
            sheet,
            min_col=min_column,
            max_col=max_column,
            min_row=min_row,
            max_row=max_row,
        )
        chart.add_data(data, titles_from_data=False)
    chart.title = operation.title
    chart.legend = chart.legend if operation.has_legend else None
    if operation.y_number_format and getattr(chart, "y_axis", None):
        chart.y_axis.numFmt = operation.y_number_format
    start_column, start_row, _, _ = range_boundaries(operation.start_cell)
    end_column, end_row, _, _ = range_boundaries(operation.end_cell)
    if end_column < start_column or end_row < start_row:
        raise ValueError("chart end_cell must be below and right of start_cell")
    from openpyxl.utils import get_column_letter

    width_pixels = sum(
        (sheet.column_dimensions[get_column_letter(column)].width or 8.43) * 7
        for column in range(start_column, end_column + 1)
    )
    height_points = sum(
        sheet.row_dimensions[row].height or 15 for row in range(start_row, end_row + 1)
    )
    chart.width = max(2.5, width_pixels / 96 * 2.54)
    chart.height = max(2.0, height_points / 72 * 2.54)
    sheet.add_chart(chart, operation.start_cell)


def _autofit(sheet: Any, range_address: str, *, rows: bool) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries

    min_column, min_row, max_column, max_row = range_boundaries(range_address)
    if rows:
        for row in range(min_row, max_row + 1):
            longest = max(
                (
                    len(str(sheet.cell(row, column).value or ""))
                    for column in range(min_column, max_column + 1)
                ),
                default=1,
            )
            sheet.row_dimensions[row].height = min(
                120, max(18, 15 * math.ceil(longest / 50))
            )
    else:
        for column in range(min_column, max_column + 1):
            longest = max(
                (
                    len(str(sheet.cell(row, column).value or ""))
                    for row in range(min_row, max_row + 1)
                ),
                default=1,
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(
                80, max(8, longest + 2)
            )


def _apply_operation(sheet: Any, operation: SheetOperation) -> None:
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
    from openpyxl.styles import PatternFill
    from openpyxl.styles.cell_style import StyleArray
    from openpyxl.worksheet.datavalidation import DataValidation as OpenpyxlValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if isinstance(operation, WriteRange):
        matrix: list[list[Any]]
        if operation.values is not None:
            matrix = operation.values
        elif operation.formulas is not None:
            matrix = operation.formulas
        else:
            matrix = [
                [datetime.fromisoformat(item) if item else None for item in row]
                for row in operation.dates or []
            ]
        _write_matrix(sheet, operation.range, matrix)
        if operation.format:
            _apply_format(sheet, operation.range, operation.format)
        return
    if isinstance(operation, StyleRange):
        _apply_format(sheet, operation.range, operation.format)
        return
    if isinstance(operation, MergeRange):
        if operation.operation == "merge":
            sheet.merge_cells(operation.range)
        else:
            sheet.unmerge_cells(operation.range)
        return
    if isinstance(operation, ClearRange):
        for cell in _iter_cells(sheet, operation.range):
            if operation.apply_to in {"contents", "all"}:
                cell.value = None
            if operation.apply_to in {"formats", "all"}:
                cell._style = StyleArray()  # noqa: SLF001 - OpenXML style reset
        return
    if isinstance(operation, FreezePanes):
        if not operation.rows and not operation.columns:
            sheet.freeze_panes = None
        else:
            from openpyxl.utils import get_column_letter

            sheet.freeze_panes = (
                f"{get_column_letter(operation.columns + 1)}{operation.rows + 1}"
            )
        return
    if isinstance(operation, DataValidation):
        escaped = [value.replace('"', '""') for value in operation.values]
        validation = OpenpyxlValidation(
            type="list", formula1=f'"{",".join(escaped)}"', allow_blank=True
        )
        sheet.add_data_validation(validation)
        validation.add(operation.range)
        return
    if isinstance(operation, ConditionalFormat):
        config = operation.config
        if operation.rule_type == "cellIs":
            rule = CellIsRule(
                operator=str(config.get("operator", "equal")),
                formula=[str(item) for item in config.get("formula", [0])],
                fill=PatternFill("solid", fgColor=_hex(config.get("fill", "#fff2cc"))),
            )
        elif operation.rule_type == "colorScale":
            rule = ColorScaleRule(
                start_type="min",
                start_color=_hex(config.get("start_color", "#f8696b")),
                mid_type="percentile",
                mid_value=50,
                mid_color=_hex(config.get("mid_color", "#ffeb84")),
                end_type="max",
                end_color=_hex(config.get("end_color", "#63be7b")),
            )
        else:
            formula = str(config.get("formula") or "TRUE")
            rule = FormulaRule(formula=[formula])
        sheet.conditional_formatting.add(operation.range, rule)
        return
    if isinstance(operation, AddTable):
        table = Table(displayName=operation.name, ref=operation.range)
        if operation.style:
            table.tableStyleInfo = TableStyleInfo(
                name=operation.style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        sheet.add_table(table)
        return
    if isinstance(operation, AddChart):
        _add_chart(sheet, operation)
        return
    if isinstance(operation, DeleteDrawings):
        sheet._charts.clear()  # noqa: SLF001 - no public bulk drawing removal
        sheet._images.clear()  # noqa: SLF001
        return
    if isinstance(operation, AutofitRange):
        _autofit(
            sheet,
            operation.range,
            rows=operation.operation == "autofit_rows",
        )
        return
    raise ValueError(f"unsupported worksheet operation: {operation.operation}")


def _open_project_workbook(project: WorkbookProject, source: Path | None) -> Any:
    from openpyxl import Workbook, load_workbook

    if source is not None:
        return load_workbook(source, data_only=False)
    workbook = Workbook()
    workbook.active.title = project.sheets[0].name
    return workbook


def _apply_project(workbook: Any, project: WorkbookProject) -> None:
    for index, plan in enumerate(project.sheets):
        if plan.name in workbook.sheetnames:
            sheet = workbook[plan.name]
        elif project.mode == "new" or plan.create_if_missing:
            if project.mode == "new" and index == 0 and len(workbook.worksheets) == 1:
                sheet = workbook.active
                sheet.title = plan.name
            else:
                sheet = workbook.create_sheet(plan.name)
        else:
            raise ValueError(f"worksheet does not exist in template: {plan.name}")
        if plan.show_grid_lines is not None:
            sheet.sheet_view.showGridLines = plan.show_grid_lines
        for operation in plan.operations:
            _apply_operation(sheet, operation)


def _formula_issues(workbook: Any) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    pattern = re.compile(_FORMULA_ERRORS)
    formula_count = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if pattern.search(value):
                        issues.append(
                            {
                                "severity": "error",
                                "code": "formula-error",
                                "message": f"{sheet.title}!{cell.coordinate} contains {value}",
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                            }
                        )
                elif isinstance(value, str) and pattern.fullmatch(value):
                    issues.append(
                        {
                            "severity": "error",
                            "code": "formula-error",
                            "message": f"{sheet.title}!{cell.coordinate} contains {value}",
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                        }
                    )
    if formula_count:
        issues.append(
            {
                "severity": "info",
                "code": "formula-recalculation-deferred",
                "message": (
                    f"{formula_count} formulas were structurally validated; "
                    "Excel will recalculate cached values when opened."
                ),
            }
        )
    return issues


def _write_manifest(workbook: Any, work_dir: Path, source: Path | None) -> Path:
    work_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = work_dir / "workbook-manifest.json"
    manifest = {
        "schemaVersion": 2,
        "engine": "evoflux-openxml",
        "sourcePath": str(source) if source else None,
        "sourceSha256": file_sha256(source) if source else None,
        "sheetCount": len(workbook.worksheets),
        "sheets": [
            {
                "name": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "mergedRanges": [str(value) for value in sheet.merged_cells.ranges],
                "tables": sorted(sheet.tables),
                "chartCount": len(sheet._charts),  # noqa: SLF001
            }
            for sheet in workbook.worksheets
        ],
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return manifest_path


def _execute_xlsx(
    action: Literal["inspect", "render", "compose"],
    *,
    project: WorkbookProject | None,
    source: Path | None,
    output: Path | None,
    work_dir: Path,
) -> XlsxPipelineResult:
    if action == "inspect":
        if source is None:
            raise ValueError("inspect requires source XLSX")
        from openpyxl import load_workbook

        workbook = load_workbook(source, data_only=False)
    else:
        if project is None:
            raise ValueError(f"{action} requires a workbook project")
        workbook = _open_project_workbook(project, source)
        _apply_project(workbook, project)
    try:
        issues = _formula_issues(workbook)
        previews = render_xlsx_workbook(workbook, work_dir / "previews")
        manifest_path = _write_manifest(workbook, work_dir, source)
        candidate: Path | None = None
        if (
            action == "compose"
            and output is not None
            and not any(issue.get("severity") == "error" for issue in issues)
        ):
            output.parent.mkdir(parents=True, exist_ok=True)
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.save(output)
            candidate = output
        return XlsxPipelineResult(
            action=action,
            work_dir=work_dir,
            source_xlsx=source,
            output=candidate,
            manifest_path=manifest_path,
            previews=previews,
            issues=issues,
            metadata={
                "engine": "evoflux-openxml",
                "sheet_count": len(workbook.worksheets),
                "sheet_names": [sheet.title for sheet in workbook.worksheets],
            },
        )
    finally:
        workbook.close()


def validate_workbook_project(
    project: WorkbookProject, source: Path | None = None
) -> dict[str, Any]:
    if project.mode == "template":
        if source is None:
            raise ValueError("template project requires source_xlsx")
        if file_sha256(source) != project.source_sha256:
            raise ValueError("source XLSX changed after inspection; inspect it again")
    elif source is not None:
        raise ValueError("new workbook project must not declare source_xlsx")
    return {
        "valid": True,
        "mode": project.mode,
        "sheet_count": len(project.sheets),
        "operation_count": sum(len(sheet.operations) for sheet in project.sheets),
        "formula_error_pattern": _FORMULA_ERRORS,
    }


async def inspect_xlsx(
    source: Path, *, workspace_root: Path, work_dir: Path
) -> XlsxPipelineResult:
    del workspace_root
    return await asyncio.to_thread(
        _execute_xlsx,
        "inspect",
        project=None,
        source=source,
        output=None,
        work_dir=work_dir,
    )


async def render_xlsx_project(
    project_path: Path, source: Path | None, *, workspace_root: Path, work_dir: Path
) -> XlsxPipelineResult:
    project = load_workbook_project(project_path)
    validate_workbook_project(project, source)
    del workspace_root
    return await asyncio.to_thread(
        _execute_xlsx,
        "render",
        project=project,
        source=source,
        output=None,
        work_dir=work_dir,
    )


async def compose_xlsx_project(
    project_path: Path,
    source: Path | None,
    output: Path,
    *,
    workspace_root: Path,
    work_dir: Path,
) -> XlsxPipelineResult:
    project = load_workbook_project(project_path)
    validate_workbook_project(project, source)
    del workspace_root
    result = await asyncio.to_thread(
        _execute_xlsx,
        "compose",
        project=project,
        source=source,
        output=output,
        work_dir=work_dir,
    )
    if not result.passed and output.exists():
        output.unlink()
        result.output = None
    return result


__all__ = [
    "WorkbookProject",
    "XlsxPipelineResult",
    "compose_xlsx_project",
    "inspect_xlsx",
    "load_workbook_project",
    "render_xlsx_project",
    "validate_workbook_project",
    "workbook_catalog",
    "xlsx_sha256",
]

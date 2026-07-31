"""Declarative, native XLSX compiler for EvoOffice."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Annotated, Any, Literal, Mapping, cast

from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
)
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator

from app.agent.builtin_skills.xlsx.scripts.qa import inspect_xlsx
from app.agent.builtin_skills.xlsx.scripts.stylekit import (
    WorkbookProfileName,
    WorkbookTheme,
    add_list_validation,
    add_status_formatting,
    add_table,
    add_variance_formatting,
    apply_workbook_profile,
    declare_content_contract,
    prepare_data_sheet,
    set_column_widths,
    style_header,
    style_input,
    style_title,
    workbook_profile,
)
from app.services.office_visual_qa_service import render_office_images

CellScalar = str | int | float | bool | date | datetime | None
CellRole = Literal["text", "input", "formula", "kpi", "section", "note"]
ChartKind = Literal["column", "bar", "line", "area", "pie", "doughnut"]


XLSX_CAPABILITIES: dict[str, dict[str, str]] = {
    "cells_and_ranges": {"create": "full", "edit": "full"},
    "formulas": {"create": "full", "edit": "full"},
    "tables_and_filters": {"create": "full", "edit": "full"},
    "charts": {"create": "full", "edit": "partial"},
    "validation_and_conditional_formatting": {"create": "full", "edit": "full"},
    "named_ranges_and_comments": {"create": "full", "edit": "full"},
    "images": {"create": "partial", "edit": "partial"},
    "pivot_tables_and_slicers": {"create": "template-first", "edit": "preserve-only"},
    "vba_and_external_queries": {"create": "unsupported", "edit": "preserve-only"},
}


class XlsxThemeSpec(BaseModel):
    model_config = ConfigDict(extra="forbid")

    font: str = Field(default="Aptos", min_length=1, max_length=80)
    ink: str = Field(default="24323D", pattern=r"^[0-9A-Fa-f]{6}$")
    muted: str = Field(default="66717C", pattern=r"^[0-9A-Fa-f]{6}$")
    accent: str = Field(default="2F6D68", pattern=r"^[0-9A-Fa-f]{6}$")
    header_fill: str = Field(default="DDE9E7", pattern=r"^[0-9A-Fa-f]{6}$")
    section_fill: str = Field(default="EEF3F2", pattern=r"^[0-9A-Fa-f]{6}$")
    input_fill: str = Field(default="FFF2CC", pattern=r"^[0-9A-Fa-f]{6}$")
    border: str = Field(default="CCD4D8", pattern=r"^[0-9A-Fa-f]{6}$")

    def to_native(self) -> WorkbookTheme:
        return WorkbookTheme(**self.model_dump())


class CellSpec(BaseModel):
    model_config = ConfigDict(extra="forbid")

    address: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]*$")
    value: CellScalar = None
    formula: str | None = None
    role: CellRole = "text"
    number_format: str | None = Field(default=None, max_length=80)
    comment: str | None = Field(default=None, max_length=1000)

    @model_validator(mode="after")
    def validate_cell(self) -> CellSpec:
        if self.formula is not None and not self.formula.startswith("="):
            raise ValueError("formula must start with '='")
        if self.formula is not None and self.value is not None:
            raise ValueError("cell cannot define both value and formula")
        if self.role == "formula" and self.formula is None:
            raise ValueError("formula role requires formula")
        return self


class TableColumnSpec(BaseModel):
    model_config = ConfigDict(extra="forbid")

    key: str = Field(min_length=1, max_length=80, pattern=r"^[A-Za-z_][A-Za-z0-9_]*$")
    label: str = Field(min_length=1, max_length=120)
    width: float = Field(default=14, ge=4, le=48)
    number_format: str | None = Field(default=None, max_length=80)
    editable: bool = False
    validation: list[str] = Field(default_factory=list, max_length=100)
    status_colors: dict[str, str] = Field(default_factory=dict)
    variance_colors: bool = False

    @field_validator("status_colors")
    @classmethod
    def validate_status_colors(cls, value: dict[str, str]) -> dict[str, str]:
        for label, color in value.items():
            if not label.strip() or not re.fullmatch(r"[0-9A-Fa-f]{6}", color):
                raise ValueError("status_colors requires non-empty labels and RGB hex")
        return value


class TableBlock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    type: Literal["table"] = "table"
    anchor: str = Field(default="A4", pattern=r"^[A-Z]{1,3}[1-9][0-9]*$")
    name: str = Field(pattern=r"^[A-Za-z_][A-Za-z0-9_]{0,254}$")
    columns: list[TableColumnSpec] = Field(min_length=1, max_length=100)
    rows: list[dict[str, CellScalar]] = Field(default_factory=list, max_length=5000)
    formula_columns: dict[str, str] = Field(default_factory=dict)
    style: str = Field(default="TableStyleMedium2", min_length=1, max_length=80)

    @model_validator(mode="after")
    def validate_table(self) -> TableBlock:
        keys = [column.key for column in self.columns]
        if len(set(keys)) != len(keys):
            raise ValueError("table column keys must be unique")
        unknown_formulas = sorted(set(self.formula_columns) - set(keys))
        if unknown_formulas:
            raise ValueError(
                "formula_columns reference unknown keys: " + ", ".join(unknown_formulas)
            )
        for key, formula in self.formula_columns.items():
            if not formula.startswith("=") or "{row}" not in formula:
                raise ValueError(
                    f"formula column {key!r} must start with '=' and contain '{{row}}'"
                )
        for index, row in enumerate(self.rows, start=1):
            unknown = sorted(set(row) - set(keys))
            if unknown:
                raise ValueError(
                    f"row {index} contains unknown keys: {', '.join(unknown)}"
                )
        return self


class KpiBlock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    type: Literal["kpi"] = "kpi"
    label_cell: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]*$")
    value_cell: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]*$")
    label: str = Field(min_length=1, max_length=120)
    value: CellScalar = None
    formula: str | None = None
    number_format: str = Field(default="#,##0", max_length=80)

    @model_validator(mode="after")
    def validate_kpi(self) -> KpiBlock:
        if (self.formula is None) == (self.value is None):
            raise ValueError("kpi requires exactly one of value or formula")
        if self.formula is not None and not self.formula.startswith("="):
            raise ValueError("kpi formula must start with '='")
        return self


class NoteBlock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    type: Literal["note"] = "note"
    range: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]*:[A-Z]{1,3}[1-9][0-9]*$")
    text: str = Field(min_length=1, max_length=1000)


class ChartBlock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    type: Literal["chart"] = "chart"
    kind: ChartKind
    source_sheet: str = Field(min_length=1, max_length=31)
    source_range: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]*:[A-Z]{1,3}[1-9][0-9]*$")
    position: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]*$")
    title: str = Field(min_length=1, max_length=160)
    width: float = Field(default=13, ge=4, le=30)
    height: float = Field(default=7.5, ge=3, le=20)
    y_number_format: str | None = Field(default=None, max_length=80)
    has_legend: bool | None = None


SheetBlock = Annotated[
    TableBlock | KpiBlock | NoteBlock | ChartBlock,
    Field(discriminator="type"),
]


class NamedRangeSpec(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str = Field(pattern=r"^[A-Za-z_][A-Za-z0-9_.]{0,254}$")
    reference: str = Field(min_length=3, max_length=240)


class WorksheetSpec(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str = Field(min_length=1, max_length=31)
    role: Literal["inputs", "data", "calculation", "summary", "tracker"] = "data"
    title: str | None = Field(default=None, max_length=180)
    title_range: str = Field(
        default="A1:H1",
        pattern=r"^[A-Z]{1,3}[1-9][0-9]*:[A-Z]{1,3}[1-9][0-9]*$",
    )
    freeze: str | None = Field(default=None, pattern=r"^[A-Z]{1,3}[1-9][0-9]*$")
    hidden: bool = False
    cells: list[CellSpec] = Field(default_factory=list, max_length=1000)
    blocks: list[SheetBlock] = Field(default_factory=list, max_length=100)
    column_widths: dict[str, float] = Field(default_factory=dict)
    landscape: bool = True

    @field_validator("column_widths")
    @classmethod
    def validate_widths(cls, value: dict[str, float]) -> dict[str, float]:
        for column, width in value.items():
            if not re.fullmatch(r"[A-Z]{1,3}", column) or not 4 <= width <= 48:
                raise ValueError("column_widths requires A1 column letters and 4..48")
        return value


class WorkbookSpec(BaseModel):
    model_config = ConfigDict(extra="forbid")

    title: str = Field(min_length=1, max_length=180)
    profile: WorkbookProfileName = "data-table"
    theme: XlsxThemeSpec = Field(default_factory=XlsxThemeSpec)
    required_fields: list[str] = Field(default_factory=list, max_length=100)
    named_ranges: list[NamedRangeSpec] = Field(default_factory=list, max_length=100)
    sheets: list[WorksheetSpec] = Field(min_length=1, max_length=40)

    @model_validator(mode="after")
    def validate_sheets(self) -> WorkbookSpec:
        names = [sheet.name.casefold() for sheet in self.sheets]
        if len(set(names)) != len(names):
            raise ValueError("worksheet names must be unique")
        if all(sheet.hidden for sheet in self.sheets):
            raise ValueError("at least one worksheet must be visible")
        available = set(names)
        for sheet in self.sheets:
            for block in sheet.blocks:
                if (
                    isinstance(block, ChartBlock)
                    and block.source_sheet.casefold() not in available
                ):
                    raise ValueError(
                        f"chart on {sheet.name!r} references missing worksheet "
                        f"{block.source_sheet!r}"
                    )
        return self


@dataclass(frozen=True)
class XlsxBuildResult:
    output: Path
    report: dict[str, Any]
    render: dict[str, Any] | None

    @property
    def passed(self) -> bool:
        return not self.report.get("errors") and not (self.render or {}).get("errors")

    def to_dict(self) -> dict[str, Any]:
        return {
            "output": str(self.output),
            "passed": self.passed,
            "report": self.report,
            "render": self.render,
        }


def workbook_catalog() -> dict[str, Any]:
    return {
        "profiles": ["data-table", "financial-model", "dashboard", "operational"],
        "sheet_roles": ["inputs", "data", "calculation", "summary", "tracker"],
        "block_types": ["table", "kpi", "note", "chart"],
        "chart_kinds": ["column", "bar", "line", "area", "pie", "doughnut"],
        "cell_roles": ["text", "input", "formula", "kpi", "section", "note"],
        "capabilities": XLSX_CAPABILITIES,
    }


def _apply_cell_style(cell, role: CellRole, theme: WorkbookTheme, profile) -> None:
    cell.font = Font(name=theme.font, size=profile.body_pt, color=theme.ink)
    cell.alignment = Alignment(vertical="center", wrap_text=role in {"note", "section"})
    if role == "input":
        style_input(cell, theme=theme)
    elif role == "formula":
        cell.font = Font(name=theme.font, size=profile.body_pt, color="008000")
    elif role == "kpi":
        cell.font = Font(
            name=theme.font,
            size=profile.kpi_pt,
            bold=True,
            color=theme.accent,
        )
    elif role == "section":
        cell.fill = PatternFill("solid", fgColor=theme.section_fill)
        cell.font = Font(
            name=theme.font,
            size=profile.header_pt,
            bold=True,
            color=theme.ink,
        )
    elif role == "note":
        cell.fill = PatternFill("solid", fgColor=theme.section_fill)
        cell.font = Font(
            name=theme.font,
            size=max(profile.body_pt - 0.5, 8),
            italic=True,
            color=theme.muted,
        )


def _render_table(sheet, block: TableBlock, *, theme: WorkbookTheme, profile) -> None:
    start_column, start_row, _, _ = range_boundaries(f"{block.anchor}:{block.anchor}")
    for offset, column in enumerate(block.columns):
        cell = sheet.cell(start_row, start_column + offset, column.label)
        sheet.column_dimensions[
            get_column_letter(start_column + offset)
        ].width = column.width
    style_header(
        sheet[start_row][start_column - 1 : start_column - 1 + len(block.columns)],
        theme=theme,
        profile=profile.name,
    )

    for row_offset, row in enumerate(block.rows, start=1):
        excel_row = start_row + row_offset
        for column_offset, column in enumerate(block.columns):
            cell = sheet.cell(excel_row, start_column + column_offset)
            formula = block.formula_columns.get(column.key)
            if formula is not None:
                cell.value = formula.format(row=excel_row)
                _apply_cell_style(cell, "formula", theme, profile)
            else:
                cell.value = row.get(column.key)
                _apply_cell_style(
                    cell,
                    "input" if column.editable else "text",
                    theme,
                    profile,
                )
            if column.number_format:
                cell.number_format = column.number_format
        sheet.row_dimensions[excel_row].height = profile.body_row_height

    end_row = start_row + max(len(block.rows), 1)
    end_column = start_column + len(block.columns) - 1
    reference = (
        f"{get_column_letter(start_column)}{start_row}:"
        f"{get_column_letter(end_column)}{end_row}"
    )
    if not block.rows:
        for column_offset in range(len(block.columns)):
            sheet.cell(start_row + 1, start_column + column_offset, None)
    add_table(sheet, reference, name=block.name, style=block.style)
    for offset, column in enumerate(block.columns):
        letter = get_column_letter(start_column + offset)
        data_range = f"{letter}{start_row + 1}:{letter}{end_row}"
        if column.validation:
            add_list_validation(sheet, data_range, column.validation)
        if column.status_colors:
            add_status_formatting(sheet, data_range, values=column.status_colors)
        if column.variance_colors:
            add_variance_formatting(sheet, data_range)


def _render_kpi(sheet, block: KpiBlock, *, theme: WorkbookTheme, profile) -> None:
    label = sheet[block.label_cell]
    label.value = block.label
    label.font = Font(
        name=theme.font,
        size=profile.header_pt,
        bold=True,
        color=theme.muted,
    )
    label.alignment = Alignment(vertical="center")
    value = sheet[block.value_cell]
    value.value = block.formula if block.formula is not None else block.value
    value.number_format = block.number_format
    _apply_cell_style(value, "kpi", theme, profile)


def _render_note(sheet, block: NoteBlock, *, theme: WorkbookTheme, profile) -> None:
    sheet.merge_cells(block.range)
    anchor = sheet[block.range.split(":")[0]]
    anchor.value = block.text
    _apply_cell_style(anchor, "note", theme, profile)
    anchor.alignment = Alignment(vertical="center", wrap_text=True)
    anchor.border = Border(
        left=Side(style="medium", color=theme.accent),
        bottom=Side(style="thin", color=theme.border),
    )


def _chart_for(kind: ChartKind):
    if kind in {"column", "bar"}:
        chart = BarChart()
        chart.type = "col" if kind == "column" else "bar"
        return chart
    return {
        "line": LineChart,
        "area": AreaChart,
        "pie": PieChart,
        "doughnut": DoughnutChart,
    }[kind]()


def _render_chart(workbook: Workbook, sheet, block: ChartBlock) -> None:
    source = workbook[block.source_sheet]
    min_col, min_row, max_col, max_row = range_boundaries(block.source_range)
    chart = _chart_for(block.kind)
    data = Reference(
        source,
        min_col=min_col + 1,
        max_col=max_col,
        min_row=min_row,
        max_row=max_row,
    )
    categories = Reference(
        source,
        min_col=min_col,
        min_row=min_row + 1,
        max_row=max_row,
    )
    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(categories)
    chart.title = block.title
    chart.style = 10
    chart.width = block.width
    chart.height = block.height
    chart.legend = None if block.has_legend is False else chart.legend
    if block.has_legend is True and chart.legend is None:
        from openpyxl.chart.legend import Legend

        chart.legend = Legend()
    if block.y_number_format and hasattr(chart, "y_axis"):
        chart.y_axis.numFmt = block.y_number_format
    if isinstance(chart, DoughnutChart):
        chart.holeSize = 62
    sheet.add_chart(chart, block.position)


def build_workbook(
    specification: WorkbookSpec | Mapping[str, Any],
    output: Path,
    *,
    render_dir: Path | None = None,
) -> XlsxBuildResult:
    """Compile a declarative workbook into an editable XLSX and validate it."""

    spec = (
        specification
        if isinstance(specification, WorkbookSpec)
        else WorkbookSpec.model_validate(specification)
    )
    output = output.expanduser().resolve()
    if output.suffix.lower() != ".xlsx":
        raise ValueError("Excel output must use the .xlsx extension")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    theme = spec.theme.to_native()
    profile = workbook_profile(spec.profile)
    apply_workbook_profile(workbook, spec.profile)
    if spec.required_fields:
        declare_content_contract(workbook, spec.required_fields)

    for sheet_spec in spec.sheets:
        workbook.create_sheet(sheet_spec.name)

    for sheet_spec in spec.sheets:
        sheet = workbook[sheet_spec.name]
        prepare_data_sheet(
            sheet,
            profile=spec.profile,
            role=sheet_spec.role,
            freeze=sheet_spec.freeze,
        )
        sheet.sheet_state = "hidden" if sheet_spec.hidden else "visible"
        sheet.page_setup.orientation = (
            "landscape" if sheet_spec.landscape else "portrait"
        )
        sheet.page_margins.left = 0.35
        sheet.page_margins.right = 0.35
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        if sheet_spec.title:
            style_title(
                sheet,
                sheet_spec.title_range,
                sheet_spec.title,
                theme=theme,
                profile=spec.profile,
            )
        if sheet_spec.column_widths:
            widths = cast(Mapping[str | int, float], sheet_spec.column_widths)
            set_column_widths(sheet, widths, profile=spec.profile)

        for cell_spec in sheet_spec.cells:
            cell = sheet[cell_spec.address]
            cell.value = (
                cell_spec.formula if cell_spec.formula is not None else cell_spec.value
            )
            _apply_cell_style(cell, cell_spec.role, theme, profile)
            if cell_spec.number_format:
                cell.number_format = cell_spec.number_format
            if cell_spec.comment:
                cell.comment = Comment(cell_spec.comment, "EvoOffice")

        charts: list[ChartBlock] = []
        for block in sheet_spec.blocks:
            if isinstance(block, TableBlock):
                _render_table(sheet, block, theme=theme, profile=profile)
            elif isinstance(block, KpiBlock):
                _render_kpi(sheet, block, theme=theme, profile=profile)
            elif isinstance(block, NoteBlock):
                _render_note(sheet, block, theme=theme, profile=profile)
            else:
                charts.append(block)
        for chart in charts:
            _render_chart(workbook, sheet, chart)

        used = sheet.calculate_dimension()
        sheet.print_area = used
        if sheet.max_row >= 2:
            sheet.print_title_rows = "1:2" if sheet_spec.title else "1:1"

    for named in spec.named_ranges:
        workbook.defined_names.add(DefinedName(named.name, attr_text=named.reference))
    workbook.properties.title = spec.title
    workbook.properties.subject = "Generated by EvoOffice XLSX Engine"
    workbook.properties.keywords = "EvoFlux, EvoOffice, editable Excel"
    workbook.save(output)

    report = inspect_xlsx(output)
    render = render_office_images(output, render_dir.resolve()) if render_dir else None
    if render and render.get("status") == "rendered":
        report["errors"].extend(render.get("errors", []))
        report["warnings"].extend(render.get("warnings", []))
    return XlsxBuildResult(output=output, report=report, render=render)


def validate_workbook(
    source: Path,
    *,
    render_dir: Path | None = None,
) -> dict[str, Any]:
    source = source.expanduser().resolve()
    report = inspect_xlsx(source)
    if render_dir:
        render = render_office_images(source, render_dir.resolve())
        report["render"] = render
        if render.get("status") == "rendered":
            report["errors"].extend(render.get("errors", []))
            report["warnings"].extend(render.get("warnings", []))
    return report


def inspect_workbook(source: Path) -> dict[str, Any]:
    """Return a compact editable-feature inventory for an existing workbook."""

    workbook = load_workbook(source, data_only=False, keep_vba=False)
    sheets = []
    for sheet in workbook.worksheets:
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "used_range": sheet.calculate_dimension(),
                "tables": list(sheet.tables),
                "charts": len(sheet._charts),
                "images": len(sheet._images),
                "validations": len(sheet.data_validations.dataValidation),
                "conditional_formats": len(sheet.conditional_formatting),
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            }
        )
    return {"sheets": sheets, "quality": inspect_xlsx(source)}


__all__ = [
    "WorkbookSpec",
    "XLSX_CAPABILITIES",
    "XlsxBuildResult",
    "build_workbook",
    "inspect_workbook",
    "validate_workbook",
    "workbook_catalog",
]

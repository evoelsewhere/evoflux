"""Deterministic, dependency-free evaluator for the XLSX preview subset.

Excel remains the authority for recalculation when a workbook is opened.  The
built-in preview engine still needs concrete display values, however, and must
not claim that a workbook passed visual QA when it only rendered formula text.
This module intentionally implements a documented, bounded subset and reports
every unsupported formula as a document-QA error.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
import math
import re
from typing import Any


_TOKEN = re.compile(
    r"\s*(?:"
    r"(?P<reference>(?:(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"\$?[A-Za-z]{1,3}\$?[1-9][0-9]*(?::\$?[A-Za-z]{1,3}\$?[1-9][0-9]*)?)|"
    r"(?P<number>(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)(?:[Ee][+-]?[0-9]+)?)|"
    r'(?P<string>"(?:[^"]|"")*")|'
    r"(?P<identifier>[A-Za-z_][A-Za-z0-9_.]*)|"
    r"(?P<operator><=|>=|<>|[+\-*/^&=<>%,():])"
    r")"
)


class FormulaEvaluationError(ValueError):
    """Raised when a formula cannot be evaluated by the built-in subset."""


@dataclass(frozen=True)
class _Token:
    kind: str
    value: str


def _tokenize(formula: str) -> list[_Token]:
    source = formula[1:] if formula.startswith("=") else formula
    tokens: list[_Token] = []
    position = 0
    while position < len(source):
        match = _TOKEN.match(source, position)
        if not match:
            raise FormulaEvaluationError(
                f"unsupported token near {source[position : position + 24]!r}"
            )
        kind = str(match.lastgroup)
        tokens.append(_Token(kind, match.group(kind)))
        position = match.end()
    tokens.append(_Token("eof", ""))
    return tokens


class _Parser:
    def __init__(self, formula: str) -> None:
        self.tokens = _tokenize(formula)
        self.index = 0

    @property
    def token(self) -> _Token:
        return self.tokens[self.index]

    def accept(self, value: str) -> bool:
        if self.token.value != value:
            return False
        self.index += 1
        return True

    def require(self, value: str) -> None:
        if not self.accept(value):
            raise FormulaEvaluationError(
                f"expected {value!r}, got {self.token.value!r}"
            )

    def parse(self) -> Any:
        node = self.comparison()
        if self.token.kind != "eof":
            raise FormulaEvaluationError(f"unexpected token {self.token.value!r}")
        return node

    def comparison(self) -> Any:
        node = self.concat()
        while self.token.value in {"=", "<>", "<", ">", "<=", ">="}:
            operator = self.token.value
            self.index += 1
            node = ("binary", operator, node, self.concat())
        return node

    def concat(self) -> Any:
        node = self.additive()
        while self.accept("&"):
            node = ("binary", "&", node, self.additive())
        return node

    def additive(self) -> Any:
        node = self.multiplicative()
        while self.token.value in {"+", "-"}:
            operator = self.token.value
            self.index += 1
            node = ("binary", operator, node, self.multiplicative())
        return node

    def multiplicative(self) -> Any:
        node = self.power()
        while self.token.value in {"*", "/"}:
            operator = self.token.value
            self.index += 1
            node = ("binary", operator, node, self.power())
        return node

    def power(self) -> Any:
        node = self.unary()
        if self.accept("^"):
            node = ("binary", "^", node, self.power())
        return node

    def unary(self) -> Any:
        if self.token.value in {"+", "-"}:
            operator = self.token.value
            self.index += 1
            return ("unary", operator, self.unary())
        node = self.primary()
        while self.accept("%"):
            node = ("percent", node)
        return node

    def primary(self) -> Any:
        token = self.token
        if self.accept("("):
            node = self.comparison()
            self.require(")")
            return node
        if token.kind == "number":
            self.index += 1
            value = float(token.value)
            return ("number", int(value) if value.is_integer() else value)
        if token.kind == "string":
            self.index += 1
            return ("string", token.value[1:-1].replace('""', '"'))
        if token.kind == "reference":
            self.index += 1
            return ("reference", token.value)
        if token.kind == "identifier":
            self.index += 1
            name = token.value.upper()
            if name in {"TRUE", "FALSE"}:
                return ("boolean", name == "TRUE")
            if not self.accept("("):
                raise FormulaEvaluationError(f"unsupported name {token.value!r}")
            arguments: list[Any] = []
            if not self.accept(")"):
                while True:
                    arguments.append(self.comparison())
                    if self.accept(")"):
                        break
                    if not (self.accept(",") or self.accept(":")):
                        raise FormulaEvaluationError(
                            f"expected function separator, got {self.token.value!r}"
                        )
            return ("function", name, arguments)
        raise FormulaEvaluationError(f"unexpected token {token.value!r}")


def _flatten(values: list[Any]) -> list[Any]:
    flattened: list[Any] = []
    for value in values:
        if isinstance(value, list):
            flattened.extend(_flatten(value))
        else:
            flattened.append(value)
    return flattened


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError as exc:
            raise FormulaEvaluationError(f"{value!r} is not numeric") from exc
    raise FormulaEvaluationError(f"{type(value).__name__} is not numeric")


def _numeric_values(values: list[Any]) -> list[float]:
    numbers: list[float] = []
    for value in _flatten(values):
        if value is None or value == "" or isinstance(value, bool):
            continue
        try:
            numbers.append(_number(value))
        except FormulaEvaluationError:
            continue
    return numbers


def _truthy(value: Any) -> bool:
    if isinstance(value, list):
        return any(_truthy(item) for item in _flatten(value))
    if isinstance(value, str):
        return value.casefold() not in {"", "false", "0"}
    return bool(value)


def _split_reference(reference: str, current_sheet: str) -> tuple[str, str]:
    if "!" not in reference:
        return current_sheet, reference.replace("$", "").upper()
    sheet_name, address = reference.rsplit("!", 1)
    if sheet_name.startswith("'") and sheet_name.endswith("'"):
        sheet_name = sheet_name[1:-1].replace("''", "'")
    return sheet_name, address.replace("$", "").upper()


@dataclass
class FormulaEvaluation:
    workbook: Any
    values: dict[tuple[str, str], Any] = field(default_factory=dict)
    issues: list[dict[str, Any]] = field(default_factory=list)
    formula_count: int = 0
    evaluated_count: int = 0
    _visiting: set[tuple[str, str]] = field(default_factory=set)
    _failed: set[tuple[str, str]] = field(default_factory=set)

    def evaluate_all(self) -> FormulaEvaluation:
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        self.formula_count += 1
                        try:
                            self.cell_value(sheet.title, cell.coordinate)
                        except FormulaEvaluationError:
                            continue
        self.evaluated_count = self.formula_count - len(self._failed)
        return self

    def cell_value(self, sheet_name: str, coordinate: str) -> Any:
        key = (sheet_name, coordinate.replace("$", "").upper())
        if key in self.values:
            return self.values[key]
        if key in self._visiting:
            self._record_failure(key, "formula-cycle", "circular reference")
            raise FormulaEvaluationError("circular reference")
        if sheet_name not in self.workbook.sheetnames:
            raise FormulaEvaluationError(f"worksheet {sheet_name!r} does not exist")
        cell = self.workbook[sheet_name][key[1]]
        value = cell.value
        if not (
            cell.data_type == "f" or (isinstance(value, str) and value.startswith("="))
        ):
            return value
        self._visiting.add(key)
        try:
            parsed = _Parser(str(value)).parse()
            computed = self._evaluate(parsed, sheet_name)
            self.values[key] = computed
            return computed
        except (FormulaEvaluationError, ZeroDivisionError, OverflowError) as exc:
            self._record_failure(key, "formula-evaluation-failed", str(exc))
            raise FormulaEvaluationError(str(exc)) from exc
        finally:
            self._visiting.discard(key)

    def display_value(self, sheet_name: str, coordinate: str, fallback: Any) -> Any:
        try:
            return self.cell_value(sheet_name, coordinate)
        except FormulaEvaluationError:
            return fallback

    def reference_values(self, reference: str, current_sheet: str) -> list[Any]:
        sheet_name, address = _split_reference(reference, current_sheet)
        if sheet_name not in self.workbook.sheetnames:
            raise FormulaEvaluationError(f"worksheet {sheet_name!r} does not exist")
        from openpyxl.utils.cell import range_boundaries

        min_column, min_row, max_column, max_row = range_boundaries(address)
        values: list[Any] = []
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                cell = self.workbook[sheet_name].cell(row, column)
                values.append(self.cell_value(sheet_name, cell.coordinate))
        return values

    def _record_failure(self, key: tuple[str, str], code: str, detail: str) -> None:
        if key in self._failed:
            return
        self._failed.add(key)
        self.issues.append(
            {
                "severity": "error",
                "code": code,
                "message": f"{key[0]}!{key[1]} could not be evaluated: {detail}",
                "sheet": key[0],
                "cell": key[1],
            }
        )

    def _evaluate(self, node: Any, current_sheet: str) -> Any:
        kind = node[0]
        if kind in {"number", "string", "boolean"}:
            return node[1]
        if kind == "reference":
            values = self.reference_values(node[1], current_sheet)
            return values if ":" in node[1] else values[0]
        if kind == "unary":
            value = _number(self._evaluate(node[2], current_sheet))
            return value if node[1] == "+" else -value
        if kind == "percent":
            return _number(self._evaluate(node[1], current_sheet)) / 100
        if kind == "binary":
            left = self._evaluate(node[2], current_sheet)
            right = self._evaluate(node[3], current_sheet)
            operator = node[1]
            if operator == "&":
                return f"{left if left is not None else ''}{right if right is not None else ''}"
            if operator in {"=", "<>", "<", ">", "<=", ">="}:
                try:
                    first: Any = _number(left)
                    second: Any = _number(right)
                except FormulaEvaluationError:
                    first = str(left or "").casefold()
                    second = str(right or "").casefold()
                return {
                    "=": first == second,
                    "<>": first != second,
                    "<": first < second,
                    ">": first > second,
                    "<=": first <= second,
                    ">=": first >= second,
                }[operator]
            first = _number(left)
            second = _number(right)
            if operator == "+":
                return first + second
            if operator == "-":
                return first - second
            if operator == "*":
                return first * second
            if operator == "/":
                if second == 0:
                    raise FormulaEvaluationError("division by zero")
                return first / second
            if operator == "^":
                return first**second
            raise FormulaEvaluationError(f"unsupported operator {operator}")
        if kind == "function":
            name = node[1]
            arguments = node[2]
            if name == "IF":
                if len(arguments) not in {2, 3}:
                    raise FormulaEvaluationError("IF expects two or three arguments")
                branch = (
                    arguments[1]
                    if _truthy(self._evaluate(arguments[0], current_sheet))
                    else (arguments[2] if len(arguments) == 3 else ("boolean", False))
                )
                return self._evaluate(branch, current_sheet)
            if name == "IFERROR":
                if len(arguments) != 2:
                    raise FormulaEvaluationError("IFERROR expects two arguments")
                try:
                    return self._evaluate(arguments[0], current_sheet)
                except FormulaEvaluationError:
                    return self._evaluate(arguments[1], current_sheet)
            values = [self._evaluate(item, current_sheet) for item in arguments]
            flattened = _flatten(values)
            numbers = _numeric_values(values)
            if name == "SUM":
                return sum(numbers)
            if name == "AVERAGE":
                if not numbers:
                    raise FormulaEvaluationError("AVERAGE has no numeric values")
                return sum(numbers) / len(numbers)
            if name == "MIN":
                if not numbers:
                    raise FormulaEvaluationError("MIN has no numeric values")
                return min(numbers)
            if name == "MAX":
                if not numbers:
                    raise FormulaEvaluationError("MAX has no numeric values")
                return max(numbers)
            if name == "COUNT":
                return len(numbers)
            if name == "COUNTA":
                return sum(value not in {None, ""} for value in flattened)
            if name == "ABS" and len(flattened) == 1:
                return abs(_number(flattened[0]))
            if name == "ROUND" and len(flattened) == 2:
                return round(_number(flattened[0]), int(_number(flattened[1])))
            if name == "AND":
                return all(_truthy(value) for value in flattened)
            if name == "OR":
                return any(_truthy(value) for value in flattened)
            if name == "NOT" and len(flattened) == 1:
                return not _truthy(flattened[0])
            raise FormulaEvaluationError(f"unsupported function {name}")
        raise FormulaEvaluationError(f"unsupported expression node {kind}")


def evaluate_workbook_formulas(workbook: Any) -> FormulaEvaluation:
    """Evaluate every formula and collect fail-closed QA evidence."""

    return FormulaEvaluation(workbook).evaluate_all()


def format_computed_value(value: Any, number_format: str = "") -> str:
    """Format a computed value for the Office-like preview surface."""

    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return "#NUM!"
        if "%" in number_format:
            decimals = 0
            if "." in number_format:
                decimals = len(number_format.split(".", 1)[1].split("%", 1)[0])
            return f"{float(value) * 100:.{decimals}f}%"
        currency = next(
            (symbol for symbol in ("$", "€", "£", "¥") if symbol in number_format),
            None,
        )
        decimals = 2 if any(token in number_format for token in (".00", ".##")) else 0
        if currency:
            return f"{currency}{float(value):,.{decimals}f}"
        if "," in number_format:
            return f"{float(value):,.{decimals}f}"
        if float(value).is_integer():
            return str(int(value))
        return f"{float(value):.10g}"
    return str(value)


__all__ = [
    "FormulaEvaluation",
    "FormulaEvaluationError",
    "evaluate_workbook_formulas",
    "format_computed_value",
]

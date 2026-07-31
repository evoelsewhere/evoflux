"""Deterministic style and calculation helpers for EvoFlux XLSX builders."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Literal, Mapping

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

WorkbookProfileName = Literal[
    "data-table",
    "financial-model",
    "dashboard",
    "operational",
]
SheetRole = Literal["inputs", "data", "calculation", "summary", "tracker"]


@dataclass(frozen=True)
class WorkbookProfile:
    """Density, typography, and navigation policy for a workbook archetype."""

    name: WorkbookProfileName
    title_pt: float
    header_pt: float
    body_pt: float
    kpi_pt: float
    title_row_height: float
    header_row_height: float
    body_row_height: float
    max_column_width: float
    zoom_scale: int
    default_freeze: str
    show_gridlines: bool


WORKBOOK_PROFILES: dict[WorkbookProfileName, WorkbookProfile] = {
    "data-table": WorkbookProfile(
        name="data-table",
        title_pt=16,
        header_pt=10,
        body_pt=9.5,
        kpi_pt=16,
        title_row_height=28,
        header_row_height=26,
        body_row_height=19,
        max_column_width=48,
        zoom_scale=90,
        default_freeze="A2",
        show_gridlines=False,
    ),
    "financial-model": WorkbookProfile(
        name="financial-model",
        title_pt=15,
        header_pt=9,
        body_pt=9,
        kpi_pt=15,
        title_row_height=26,
        header_row_height=22,
        body_row_height=18,
        max_column_width=36,
        zoom_scale=85,
        default_freeze="B5",
        show_gridlines=False,
    ),
    "dashboard": WorkbookProfile(
        name="dashboard",
        title_pt=20,
        header_pt=10,
        body_pt=9.5,
        kpi_pt=18,
        title_row_height=32,
        header_row_height=24,
        body_row_height=20,
        max_column_width=42,
        zoom_scale=90,
        default_freeze="A1",
        show_gridlines=False,
    ),
    "operational": WorkbookProfile(
        name="operational",
        title_pt=15,
        header_pt=9,
        body_pt=8.5,
        kpi_pt=14,
        title_row_height=26,
        header_row_height=24,
        body_row_height=18,
        max_column_width=36,
        zoom_scale=85,
        default_freeze="A5",
        show_gridlines=False,
    ),
}


def workbook_profile(
    name: WorkbookProfileName = "data-table",
) -> WorkbookProfile:
    return WORKBOOK_PROFILES[name]


@dataclass(frozen=True)
class WorkbookTheme:
    """A neutral analytical workbook theme."""

    font: str = "Aptos"
    ink: str = "24323D"
    muted: str = "66717C"
    accent: str = "2F6D68"
    header_fill: str = "DDE9E7"
    section_fill: str = "EEF3F2"
    input_fill: str = "FFF2CC"
    border: str = "CCD4D8"


def apply_workbook_profile(
    workbook: Workbook,
    name: WorkbookProfileName,
) -> None:
    """Persist a non-visual profile marker for QA and downstream editing."""

    marker_name = "_EVOFLUX_PROFILE"
    if marker_name in workbook.defined_names:
        del workbook.defined_names[marker_name]
    workbook.defined_names.add(
        DefinedName(marker_name, attr_text=f'"{name}"', hidden=True)
    )
    configure_workbook(workbook)


def declare_content_contract(
    workbook: Workbook,
    required_fields: list[str],
) -> None:
    """Persist required semantic fields so QA can detect silent content loss."""

    cleaned = [field.strip() for field in required_fields if field.strip()]
    if not cleaned:
        raise ValueError("required_fields must not be empty")
    marker_name = "_EVOFLUX_REQUIRED_FIELDS"
    if marker_name in workbook.defined_names:
        del workbook.defined_names[marker_name]
    payload = "|".join(dict.fromkeys(cleaned)).replace('"', '""')
    workbook.defined_names.add(
        DefinedName(marker_name, attr_text=f'"{payload}"', hidden=True)
    )


def configure_workbook(
    workbook: Workbook,
    *,
    profile: WorkbookProfileName | None = None,
) -> None:
    """Request a full automatic calculation when Excel opens the file."""
    calculation = workbook.calculation
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True
    if profile is not None:
        apply_workbook_profile(workbook, profile)


def style_title(
    sheet,
    cell_range: str,
    title: str,
    *,
    theme: WorkbookTheme = WorkbookTheme(),
    profile: WorkbookProfileName = "data-table",
) -> None:
    policy = workbook_profile(profile)
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = title
    cell.font = Font(
        name=theme.font,
        size=policy.title_pt,
        bold=True,
        color=theme.ink,
    )
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[cell.row].height = policy.title_row_height


def style_header(
    cells,
    *,
    theme: WorkbookTheme = WorkbookTheme(),
    profile: WorkbookProfileName = "data-table",
) -> None:
    policy = workbook_profile(profile)
    thin = Side(style="thin", color=theme.border)
    for cell in cells:
        cell.font = Font(
            name=theme.font,
            size=policy.header_pt,
            bold=True,
            color=theme.ink,
        )
        cell.fill = PatternFill("solid", fgColor=theme.header_fill)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=thin)


def style_input(cell, *, theme: WorkbookTheme = WorkbookTheme()) -> None:
    cell.font = Font(name=theme.font, color="0000FF")
    cell.fill = PatternFill("solid", fgColor=theme.input_fill)


def style_body_range(
    sheet,
    cell_range: str,
    *,
    profile: WorkbookProfileName = "data-table",
    theme: WorkbookTheme = WorkbookTheme(),
) -> None:
    """Apply compact body typography without changing values or number formats."""

    policy = workbook_profile(profile)
    for row in sheet[cell_range]:
        for cell in row:
            cell.font = Font(
                name=theme.font,
                size=policy.body_pt,
                color=theme.ink,
            )
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center",
                wrap_text=cell.alignment.wrap_text,
            )
        if row:
            row_number = row[0].row
            if not sheet.row_dimensions[row_number].height:
                sheet.row_dimensions[row_number].height = policy.body_row_height


def set_column_widths(
    sheet,
    widths: Mapping[str | int, float],
    *,
    profile: WorkbookProfileName = "data-table",
) -> None:
    maximum = workbook_profile(profile).max_column_width
    for key, width in widths.items():
        letter = get_column_letter(key) if isinstance(key, int) else key
        sheet.column_dimensions[letter].width = min(max(width, 4), maximum)


def add_table(
    sheet,
    reference: str,
    *,
    name: str,
    style: str = "TableStyleMedium2",
) -> Table:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    return table


def prepare_data_sheet(
    sheet,
    *,
    profile: WorkbookProfileName = "data-table",
    role: SheetRole = "data",
    freeze: str | None = None,
    auto_filter: str | None = None,
) -> None:
    policy = workbook_profile(profile)
    sheet.freeze_panes = freeze or policy.default_freeze
    sheet.sheet_view.showGridLines = policy.show_gridlines
    sheet.sheet_view.zoomScale = policy.zoom_scale
    if auto_filter:
        sheet.auto_filter.ref = auto_filter
    sheet.sheet_properties.tabColor = {
        "inputs": "F4B183",
        "data": "5B9BD5",
        "calculation": "A5A5A5",
        "summary": "70AD47",
        "tracker": "4472C4",
    }[role]
    sheet.print_options.horizontalCentered = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def add_list_validation(
    sheet,
    cell_range: str,
    values: list[str],
    *,
    allow_blank: bool = True,
) -> DataValidation:
    """Add editable categorical validation with a visible dropdown."""

    if not values:
        raise ValueError("values must not be empty")
    escaped = [value.replace('"', '""') for value in values]
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(escaped)}"',
        allow_blank=allow_blank,
    )
    validation.error = "Choose a value from the list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Select an allowed value."
    validation.promptTitle = "Editable field"
    sheet.add_data_validation(validation)
    validation.add(cell_range)
    return validation


def add_status_formatting(
    sheet,
    cell_range: str,
    *,
    values: dict[str, str] | None = None,
) -> None:
    """Apply formula-driven status colors that react to future edits."""

    palette = values or {
        "On track": "E2F0D9",
        "At risk": "FFF2CC",
        "Blocked": "FCE4D6",
        "Done": "DDEBF7",
    }
    anchor = cell_range.split(":")[0]
    for value, fill in palette.items():
        sheet.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'{anchor}="{value}"'],
                fill=PatternFill("solid", fgColor=fill),
            ),
        )


def add_variance_formatting(sheet, cell_range: str) -> None:
    """Use native conditional formatting for positive/negative variance."""

    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor="FCE4D6"),
        ),
    )
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor="E2F0D9"),
        ),
    )

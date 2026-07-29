"""Structural, formula, chart, and optional render QA for XLSX files."""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_FORMULA
from openpyxl.utils import get_column_letter

for parent in Path(__file__).resolve().parents:
    if (parent / "app" / "services").is_dir():
        sys.path.insert(0, str(parent))
        break

from app.services.office_visual_qa_service import (  # noqa: E402
    compare_rendered_images,
    render_office_images,
)
from app.agent.builtin_skills.xlsx.scripts.stylekit import (  # noqa: E402
    WorkbookProfileName,
    workbook_profile,
)

PLACEHOLDER_RE = re.compile(
    r"(\{\{[^{}]+\}\}|<TODO>|<PLACEHOLDER>|lorem ipsum|\bxxxx\b)",
    re.IGNORECASE,
)
FORMULA_ERROR_RE = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|NULL!|NUM!|N/A)")
IMPLICIT_ARRAY_RE = re.compile(
    r"\bMATCH\s*\(\s*TRUE\s*\(\s*\)\s*,[^)]*[<>=][^)]*,",
    re.IGNORECASE,
)


def _render(source: Path, render_dir: Path) -> dict[str, Any]:
    return render_office_images(source, render_dir)


def _defined_text(workbook, name: str) -> str | None:
    defined = workbook.defined_names.get(name)
    if defined is None:
        return None
    value = str(defined.attr_text or "")
    if len(value) >= 2 and value.startswith('"') and value.endswith('"'):
        return value[1:-1].replace('""', '"')
    return value or None


def _workbook_profile(workbook) -> WorkbookProfileName:
    value = _defined_text(workbook, "_EVOFLUX_PROFILE")
    if value in {"data-table", "financial-model", "dashboard", "operational"}:
        return cast(WorkbookProfileName, value)
    return "data-table"


def _required_fields(workbook) -> list[str]:
    value = _defined_text(workbook, "_EVOFLUX_REQUIRED_FIELDS")
    if not value:
        return []
    return [field.strip() for field in value.split("|") if field.strip()]


def _chart_formulae(chart) -> list[str]:
    references: list[str] = []
    for series in chart.series:
        for attribute in ("val", "cat", "xVal", "yVal"):
            value = getattr(series, attribute, None)
            if value is None:
                continue
            for reference_name in ("numRef", "strRef"):
                reference = getattr(value, reference_name, None)
                formula = getattr(reference, "f", None)
                if formula:
                    references.append(formula)
    return references


def inspect_xlsx(source: Path) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    try:
        with zipfile.ZipFile(source) as package:
            bad_entry = package.testzip()
            if bad_entry:
                errors.append(f"Corrupt ZIP entry: {bad_entry}")
            names = set(package.namelist())
            for required in (
                "[Content_Types].xml",
                "_rels/.rels",
                "xl/workbook.xml",
                "xl/styles.xml",
            ):
                if required not in names:
                    errors.append(f"Missing package part: {required}")
    except (OSError, zipfile.BadZipFile) as exc:
        return {"errors": [f"Invalid XLSX package: {exc}"], "warnings": []}

    keep_vba = source.suffix.lower() == ".xlsm"
    try:
        workbook = load_workbook(source, data_only=False, keep_vba=keep_vba)
        cached = load_workbook(source, data_only=True, keep_vba=keep_vba)
    except Exception as exc:
        return {
            "errors": [f"openpyxl could not open the package: {exc}"],
            "warnings": [],
        }

    if not workbook.sheetnames:
        errors.append("Workbook contains no worksheets")
    profile_name = _workbook_profile(workbook)
    policy = workbook_profile(profile_name)
    required_fields = _required_fields(workbook)
    formula_count = 0
    chart_count = 0
    visible_labels: set[str] = set()
    sheet_metrics: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        if (
            len(workbook.worksheets) > 1
            and sheet.title == "Sheet"
            and sheet.max_row == 1
            and sheet.max_column == 1
            and sheet["A1"].value is None
        ):
            errors.append("Blank default worksheet 'Sheet' was left in the workbook")

        cached_sheet = cached[sheet.title]
        sheet_formula_count = 0
        nonempty_cells = 0
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is not None:
                    nonempty_cells += 1
                if isinstance(value, str) and cell.data_type != TYPE_FORMULA:
                    visible_labels.add(value.strip().casefold())
                if isinstance(value, str) and PLACEHOLDER_RE.search(value):
                    errors.append(
                        f"{sheet.title}!{cell.coordinate}: unresolved placeholder"
                    )
                if cell.data_type == TYPE_FORMULA:
                    formula_count += 1
                    sheet_formula_count += 1
                    formula = str(value)
                    if FORMULA_ERROR_RE.search(formula):
                        errors.append(
                            f"{sheet.title}!{cell.coordinate}: broken formula reference"
                        )
                    if IMPLICIT_ARRAY_RE.search(formula):
                        warnings.append(
                            f"{sheet.title}!{cell.coordinate}: implicit array formula "
                            "may not work in older Excel versions"
                        )
                    cached_value = cached_sheet[cell.coordinate].value
                    if isinstance(cached_value, str) and FORMULA_ERROR_RE.search(
                        cached_value
                    ):
                        errors.append(
                            f"{sheet.title}!{cell.coordinate}: cached formula error "
                            f"{cached_value}"
                        )

        for chart in sheet._charts:
            chart_count += 1
            formulae = _chart_formulae(chart)
            if not chart.series or not formulae:
                errors.append(f"{sheet.title}: a chart has no usable data series")
            for formula in formulae:
                if "#REF!" in formula:
                    errors.append(f"{sheet.title}: chart contains a broken reference")

        oversized_columns = []
        for column in range(1, min(sheet.max_column, 200) + 1):
            letter = get_column_letter(column)
            width = sheet.column_dimensions[letter].width
            if width is not None and width > policy.max_column_width:
                warnings.append(
                    f"{sheet.title}!{letter}: {width:.1f}-character column exceeds "
                    f"the {profile_name} profile cap of {policy.max_column_width:.1f}"
                )
                oversized_columns.append(letter)

        oversized_rows = [
            row_number
            for row_number, dimension in sheet.row_dimensions.items()
            if dimension.height is not None and dimension.height > 60
        ]
        if oversized_rows:
            warnings.append(
                f"{sheet.title}: {len(oversized_rows)} row(s) exceed 60pt; "
                "check for accidental deep wrapping or oversized KPI blocks"
            )

        table_count = len(sheet.tables)
        validation_count = len(sheet.data_validations.dataValidation)
        conditional_format_count = len(sheet.conditional_formatting)
        has_filter = bool(sheet.auto_filter.ref) or table_count > 0
        if (
            profile_name in {"data-table", "operational"}
            and sheet.max_row >= 25
            and sheet.max_column >= 3
            and not has_filter
        ):
            warnings.append(
                f"{sheet.title}: {sheet.max_row} rows have no Excel Table or filter"
            )
        if (
            profile_name in {"data-table", "financial-model", "operational"}
            and sheet.max_row >= 25
            and sheet.freeze_panes is None
        ):
            warnings.append(
                f"{sheet.title}: long sheet has no frozen navigation headers"
            )
        if (
            profile_name == "operational"
            and sheet.max_row >= 10
            and validation_count == 0
        ):
            warnings.append(
                f"{sheet.title}: operational tracker has no editable data validation"
            )
        if (
            profile_name == "operational"
            and sheet.max_row >= 10
            and conditional_format_count == 0
        ):
            warnings.append(
                f"{sheet.title}: operational tracker has no reactive conditional "
                "formatting for status, risk, or variance"
            )

        sheet_metrics.append(
            {
                "sheet": sheet.title,
                "used_range": f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
                "nonempty_cells": nonempty_cells,
                "formulas": sheet_formula_count,
                "tables": table_count,
                "charts": len(sheet._charts),
                "merged_ranges": len(sheet.merged_cells.ranges),
                "validations": validation_count,
                "conditional_formats": conditional_format_count,
                "freeze_panes": (
                    str(sheet.freeze_panes) if sheet.freeze_panes is not None else None
                ),
                "auto_filter": sheet.auto_filter.ref,
                "oversized_columns": oversized_columns,
                "oversized_rows": oversized_rows,
            }
        )

    calculation = workbook.calculation
    if calculation.calcMode not in (None, "auto"):
        warnings.append(
            f"Workbook calculation mode is {calculation.calcMode!r}, not auto"
        )
    missing_fields = [
        field for field in required_fields if field.casefold() not in visible_labels
    ]
    if missing_fields:
        errors.append(
            "Content contract is missing required fields: " + ", ".join(missing_fields)
        )
    return {
        "errors": errors,
        "warnings": warnings,
        "sheets": workbook.sheetnames,
        "profile": profile_name,
        "required_fields": required_fields,
        "sheet_metrics": sheet_metrics,
        "formulas": formula_count,
        "charts": chart_count,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("file", type=Path)
    parser.add_argument("--render-dir", type=Path)
    parser.add_argument("--compare-to", type=Path)
    args = parser.parse_args()

    report = inspect_xlsx(args.file)
    if args.render_dir:
        report["render"] = _render(args.file, args.render_dir)
        render = report["render"]
        if render["status"] == "rendered":
            report["errors"].extend(render["errors"])
            report["warnings"].extend(render["warnings"])
            if args.compare_to:
                reference_dir = args.render_dir / "reference"
                reference = _render(args.compare_to, reference_dir)
                report["visual_diff"] = compare_rendered_images(
                    list(reference.get("images", [])),
                    list(render.get("images", [])),
                )
                report["errors"].extend(report["visual_diff"]["errors"])
        else:
            report["warnings"].append(
                f"Visual QA {render['status']}: {render.get('reason', 'unknown reason')}"
            )
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 1 if report["errors"] else 0


if __name__ == "__main__":
    sys.exit(main())
